import calendar
from datetime import datetime
from io import BytesIO

from django.utils import timezone
from incidents.models import Incident
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                TableStyle)


class MonthlyReportService:
    """
    Generate monthly SOC incident reports.
    """

    @staticmethod
    def get_previous_month_range():
        """
        Return timezone-aware start/end datetimes
        for the previous calendar month.
        """

        now = timezone.localtime()

        if now.month == 1:
            year = now.year - 1
            month = 12
        else:
            year = now.year
            month = now.month - 1

        last_day = calendar.monthrange(
            year,
            month,
        )[1]

        current_timezone = timezone.get_current_timezone()

        start = timezone.make_aware(
            datetime(
                year,
                month,
                1,
                0,
                0,
                0,
            ),
            current_timezone,
        )

        end = timezone.make_aware(
            datetime(
                year,
                month,
                last_day,
                23,
                59,
                59,
                999999,
            ),
            current_timezone,
        )

        return start, end

    @staticmethod
    def get_incidents(start, end):
        """
        Get incidents created inside the requested range.
        """

        return (
            Incident.objects
            .filter(
                created_at__gte=start,
                created_at__lte=end,
            )
            .select_related(
                "assigned_to",
                "created_by",
            )
            .order_by("-created_at")
        )

    @classmethod
    def generate_pdf(cls, start=None, end=None):
        """
        Generate previous month's PDF report.
        """

        if start is None or end is None:
            start, end = (
                cls.get_previous_month_range()
            )

        incidents = cls.get_incidents(
            start,
            end,
        )

        buffer = BytesIO()

        document = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=15 * mm,
            leftMargin=15 * mm,
            topMargin=15 * mm,
            bottomMargin=15 * mm,
            title="AI SOC Analyst Monthly Report",
            author="AI SOC Analyst",
        )

        styles = getSampleStyleSheet()

        story = [
            Paragraph(
                "AI SOC Analyst — Monthly Incident Report",
                styles["Title"],
            ),
            Spacer(1, 5 * mm),
            Paragraph(
                (
                    f"Period: "
                    f"{start.strftime('%d %B %Y')} - "
                    f"{end.strftime('%d %B %Y')}"
                ),
                styles["Normal"],
            ),
            Spacer(1, 3 * mm),
            Paragraph(
                f"Total Incidents: {incidents.count()}",
                styles["Heading2"],
            ),
            Spacer(1, 5 * mm),
        ]

        data = [
            [
                "Incident ID",
                "Title",
                "Severity",
                "Status",
                "Assigned To",
                "Created At",
            ]
        ]

        for incident in incidents:
            assigned_to = (
                str(incident.assigned_to)
                if incident.assigned_to
                else "-"
            )

            data.append(
                [
                    incident.incident_id,
                    Paragraph(
                        incident.title,
                        styles["BodyText"],
                    ),
                    incident.severity,
                    incident.status,
                    assigned_to,
                    timezone.localtime(
                        incident.created_at
                    ).strftime(
                        "%d-%m-%Y %H:%M"
                    ),
                ]
            )

        table = Table(
            data,
            repeatRows=1,
            colWidths=[
                30 * mm,
                65 * mm,
                25 * mm,
                30 * mm,
                45 * mm,
                40 * mm,
            ],
        )

        table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.lightgrey,
                    ),
                    (
                        "FONTNAME",
                        (0, 0),
                        (-1, 0),
                        "Helvetica-Bold",
                    ),
                    (
                        "GRID",
                        (0, 0),
                        (-1, -1),
                        0.5,
                        colors.grey,
                    ),
                    (
                        "VALIGN",
                        (0, 0),
                        (-1, -1),
                        "MIDDLE",
                    ),
                    (
                        "TOPPADDING",
                        (0, 0),
                        (-1, -1),
                        6,
                    ),
                    (
                        "BOTTOMPADDING",
                        (0, 0),
                        (-1, -1),
                        6,
                    ),
                ]
            )
        )

        story.append(table)

        document.build(story)

        buffer.seek(0)

        return buffer

    @classmethod
    def generate_excel(
        cls,
        start=None,
        end=None,
    ):
        """
        Generate previous month's Excel report.
        """

        if start is None or end is None:
            start, end = (
                cls.get_previous_month_range()
            )

        incidents = cls.get_incidents(
            start,
            end,
        )

        workbook = Workbook()

        worksheet = workbook.active
        worksheet.title = "Monthly Incidents"

        worksheet.merge_cells("A1:H1")

        worksheet["A1"] = (
            "AI SOC Analyst - Monthly Incident Report"
        )

        worksheet["A1"].font = Font(
            bold=True,
            size=16,
        )

        worksheet["A1"].alignment = Alignment(
            horizontal="center"
        )

        worksheet.merge_cells("A2:H2")

        worksheet["A2"] = (
            f"Period: "
            f"{start.strftime('%d %B %Y')} - "
            f"{end.strftime('%d %B %Y')}"
        )

        worksheet["A2"].alignment = Alignment(
            horizontal="center"
        )

        worksheet["A4"] = (
            f"Total Incidents: {incidents.count()}"
        )

        worksheet["A4"].font = Font(
            bold=True
        )

        headers = [
            "Incident ID",
            "Title",
            "Description",
            "Severity",
            "Status",
            "Assigned To",
            "Created By",
            "Created At",
        ]

        header_row = 6

        for column, header in enumerate(
            headers,
            start=1,
        ):
            cell = worksheet.cell(
                row=header_row,
                column=column,
                value=header,
            )

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center"
            )

        row = 7

        for incident in incidents:
            assigned_to = (
                str(incident.assigned_to)
                if incident.assigned_to
                else "-"
            )

            created_by = (
                str(incident.created_by)
                if incident.created_by
                else "-"
            )

            created_at = timezone.localtime(
                incident.created_at
            ).replace(
                tzinfo=None
            )

            values = [
                incident.incident_id,
                incident.title,
                incident.description,
                incident.severity,
                incident.status,
                assigned_to,
                created_by,
                created_at,
            ]

            for column, value in enumerate(
                values,
                start=1,
            ):
                cell = worksheet.cell(
                    row=row,
                    column=column,
                    value=value,
                )

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

            worksheet.cell(
                row=row,
                column=8,
            ).number_format = (
                "dd-mm-yyyy hh:mm"
            )

            row += 1

        widths = [
            20,
            35,
            60,
            15,
            18,
            25,
            25,
            22,
        ]

        for index, width in enumerate(
            widths,
            start=1,
        ):
            worksheet.column_dimensions[
                get_column_letter(index)
            ].width = width

        worksheet.freeze_panes = "A7"

        if row > 7:
            worksheet.auto_filter.ref = (
                f"A6:H{row - 1}"
            )
        else:
            worksheet.auto_filter.ref = "A6:H6"

        buffer = BytesIO()

        workbook.save(buffer)

        buffer.seek(0)

        return buffer