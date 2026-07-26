from io import BytesIO

from django.utils import timezone
from incidents.models import Incident
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class ExcelReportService:
    """
    Service responsible for generating SOC incident
    reports in Excel format.
    """

    @staticmethod
    def generate_incident_report():
        """
        Generate an Excel workbook containing
        all incidents.
        """

        workbook = Workbook()

        worksheet = workbook.active
        worksheet.title = "Incident Report"

        # --------------------------------------------------
        # Report title
        # --------------------------------------------------

        worksheet.merge_cells(
            "A1:H1"
        )

        title_cell = worksheet["A1"]

        title_cell.value = (
            "AI SOC Analyst - Incident Report"
        )

        title_cell.font = Font(
            bold=True,
            size=16,
        )

        title_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        worksheet.row_dimensions[1].height = 25

        # --------------------------------------------------
        # Generated date
        # --------------------------------------------------

        generated_at = timezone.localtime()

        worksheet.merge_cells(
            "A2:H2"
        )

        generated_cell = worksheet["A2"]

        generated_cell.value = (
            "Generated: "
            + generated_at.strftime(
                "%d %B %Y %H:%M:%S"
            )
        )

        generated_cell.alignment = Alignment(
            horizontal="center"
        )

        # --------------------------------------------------
        # Incident queryset
        # --------------------------------------------------

        incidents = (
            Incident.objects
            .select_related(
                "assigned_to",
                "created_by",
            )
            .order_by("-created_at")
        )

        worksheet["A4"] = (
            f"Total Incidents: {incidents.count()}"
        )

        worksheet["A4"].font = Font(
            bold=True
        )

        # --------------------------------------------------
        # Headers
        # --------------------------------------------------

        headers = [
            "Incident ID",
            "Title",
            "Description",
            "Severity",
            "Status",
            "Assigned To",
            "Created By",
            "Created At",
        ]

        header_row = 6

        for column_number, header in enumerate(
            headers,
            start=1,
        ):
            cell = worksheet.cell(
                row=header_row,
                column=column_number,
                value=header,
            )

            cell.font = Font(
                bold=True
            )

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="D9EAF7",
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        # --------------------------------------------------
        # Incident data
        # --------------------------------------------------

        data_row = header_row + 1

        for incident in incidents:
            assigned_to = "-"

            if incident.assigned_to:
                assigned_to = str(
                    incident.assigned_to
                )

            created_by = "-"

            if incident.created_by:
                created_by = str(
                    incident.created_by
                )

            created_at = timezone.localtime(
                incident.created_at
            ).replace(
                tzinfo=None
            )

            values = [
                incident.incident_id,
                incident.title,
                incident.description,
                incident.severity,
                incident.status,
                assigned_to,
                created_by,
                created_at,
            ]

            for column_number, value in enumerate(
                values,
                start=1,
            ):
                cell = worksheet.cell(
                    row=data_row,
                    column=column_number,
                    value=value,
                )

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

            # Excel date format
            worksheet.cell(
                row=data_row,
                column=8,
            ).number_format = (
                "dd-mm-yyyy hh:mm"
            )

            data_row += 1

        # --------------------------------------------------
        # Column widths
        # --------------------------------------------------

        column_widths = {
            1: 20,
            2: 35,
            3: 60,
            4: 15,
            5: 18,
            6: 25,
            7: 25,
            8: 22,
        }

        for column_number, width in (
            column_widths.items()
        ):
            column_letter = get_column_letter(
                column_number
            )

            worksheet.column_dimensions[
                column_letter
            ].width = width

        # --------------------------------------------------
        # Freeze header
        # --------------------------------------------------

        worksheet.freeze_panes = "A7"

        # --------------------------------------------------
        # Auto filter
        # --------------------------------------------------

        if data_row > header_row + 1:
            worksheet.auto_filter.ref = (
                f"A{header_row}:H{data_row - 1}"
            )
        else:
            worksheet.auto_filter.ref = (
                f"A{header_row}:H{header_row}"
            )

        # --------------------------------------------------
        # Save workbook to memory
        # --------------------------------------------------

        buffer = BytesIO()

        workbook.save(buffer)

        buffer.seek(0)

        return buffer